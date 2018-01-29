from openpyxl import Workbook

from flask import Blueprint, send_from_directory
from flask_jwt_extended import get_jwt_identity, jwt_required
from flask_restful import Api, abort
from flasgger import swag_from

from app.docs.admin.apply.goingout import *
from app.models.account import AdminModel, StudentModel
from app.views import BaseResource

from utils.apply_excel_manager import get_cells, ready_worksheet

api = Api(Blueprint('admin-goingout-api', __name__))
api.prefix = '/admin'


@api.resource('/goingout')
class GoingoutDownload(BaseResource):
    @swag_from(GOINGOUT_DOWNLOAD_GET)
    @jwt_required
    def get(self):
        """
        외출신청 엑셀 다운로드
        """
        admin = AdminModel.objects(id=get_jwt_identity()).first()
        if not admin:
            abort(403)

        wb = Workbook()
        ws = wb.active

        ready_worksheet(ws)

        for student in StudentModel.objects:
            number_cell, name_cell, status_cell = get_cells(student)

            ws[number_cell] = student.number
            ws[name_cell] = student.name

            goingout_apply = student.goingout_apply

            if goingout_apply.on_saturday and goingout_apply.on_sunday:
                status = '토요일, 일요일 외출'
            elif goingout_apply.on_saturday:
                status = '토요일 외출'
            elif goingout_apply.on_sunday:
                status = '일요일 외출'
            else:
                status = ''

            ws[status_cell] = status

        wb.save('goingout.xlsx')
        wb.close()

        return send_from_directory('../', 'goingout.xlsx')
